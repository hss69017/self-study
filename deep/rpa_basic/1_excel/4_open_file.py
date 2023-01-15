from openpyxl import load_workbook # 파일 불러오기
wb = load_workbook("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample.xlsx")
ws = wb.active

# cell 데이터 불러오기
# for x in range(1, 11):
#     for y in range(1,11):
#         print(ws.cell(row = x, column = y).value, end = " ") # end="~" 이것은 뒤에 ~를 붙여서 출력
#     print()

# cell 개수를 모를 때
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row = x, column = y).value, end = " ")
    print()