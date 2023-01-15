from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 최초 데이터 입력
ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])
ws.append([1,10,8,5,14,26,12])
ws.append([2,7,3,7,15,24,18])
ws.append([3,9,5,8,8,12,4])
ws.append([4,7,8,7,17,21,18])
ws.append([5,7,8,7,16,25,15])
ws.append([6,3,5,8,8,17,0])
ws.append([7,4,9,10,16,27,18])
ws.append([8,6,6,6,15,19,17])
ws.append([9,10,10,9,19,30,19])
ws.append([10,9,8,8,20,25,20])

# 퀴즈2 10으로 수정
for x in range(2, ws.max_row + 1):
    ws.cell(row = x, column = 4).value = 10

# 총점 추가
ws.cell(row = 1, column = ws.max_column + 1, value = "총점")
for x in range(2, ws.max_row + 1):
    ws.cell(row = x, column = ws.max_column, value = "=SUM(B{}:G{})".format(x, x))

# 성적 추가
ws.cell(row = 1, column = ws.max_column + 1, value = "성적")
for x in range(2, ws.max_row + 1):
    total = 0
    for y in range(2, 8):
        total += int(ws.cell(row = x, column = y).value)

    if int(ws["B{}".format(x)].value) < 5:
        ws.cell(row = x, column = ws.max_column, value = "F")
    elif total >= 90:
        ws.cell(row = x, column = ws.max_column, value = "A")
    elif total >= 80:
        ws.cell(row = x, column = ws.max_column, value = "B")
    elif total >= 70:
        ws.cell(row = x, column = ws.max_column, value = "C")
    else:
        ws.cell(row = x, column = ws.max_column, value = "D")

wb.save("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/scores.xlsx")