from openpyxl import load_workbook
wb = load_workbook("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample.xlsx")
ws = wb.active

# ws.move_range("B1:C11", rows = 0, cols = 1) # 해당 셀 영역을 row는 ~만큼, cols는 ~만큼 이동
# ws["B1"].value = "국어"

ws.move_range("C1:C11", rows = 5, cols = -1)

wb.save("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample_korean.xlsx")