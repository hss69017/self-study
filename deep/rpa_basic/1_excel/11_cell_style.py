from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

ws.column_dimensions["A"].width = 5 # A열 너비 5로 설정
ws.row_dimensions[1].height = 50 # 1행 높이 50으로 설정

a1.font = Font(color = "FF0000", italic = True, bold = True) # 글자 색상 빨강, 이탤릭, 두껍게
b1.font = Font(color = "CC33FF", name = "Arial", strike = True) # 폰트는 Arial, 취소선
c1.font = Font(color = "0000FF", size = 20, underline = "single") # 글자크기 20, 밑줄 1줄(두 줄 하려면 double)

# 테두리 적용
thin_border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 90점 넘는 셀에 대해서 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 셀 정렬 / left, right, top, bottom 가능
        cell.alignment = Alignment(horizontal = "center", vertical = "center")

        if cell.column == 1:
            continue
        
        # 셀이 정수형 데이터이고 90점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor = "00FF00", fill_type = "solid")
            cell.font = Font(color = "FF0000")

# 틀 고정
ws.freeze_panes = "B2"

wb.save("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample_style.xlsx")