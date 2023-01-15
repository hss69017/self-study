from openpyxl import load_workbook
wb = load_workbook("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample_merge.xlsx")
ws = wb.active

# 셀 병합 해제
ws.unmerge_cells("B2:D2")
wb.save("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample_unmerge.xlsx")