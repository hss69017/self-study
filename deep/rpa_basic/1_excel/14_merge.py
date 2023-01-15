from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 병합하기
ws.merge_cells("B2:D2") # B2 ~ D2 셀 병합
ws["B2"].value = "Merged Cell"

wb.save("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample_merge.xlsx")