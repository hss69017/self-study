import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today() # 오늘 날짜 정보
ws["A2"] = "=SUM(1, 2, 3)"
ws["A3"] = "=AVERAGE(1, 2, 3)"

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"

wb.save("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample_formula.xlsx")