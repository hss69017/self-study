from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet를 기본 이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff00ff" # RGB 형태로 값을 넣어주면 탭 색상 변경 / https://www.w3schools.com/colors/colors_rgb.asp

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 sheet 생성

new_ws = wb["NewSheet"] # 파일에 주어진 이름의 new_ws sheet를 생성(dict 형태)

print(wb.sheetnames) # 모든 sheet 이름 확인

# sheet 복사
new_ws["A1"] = "Test" # new_ws sheet의 A1 칸에 Test 입력
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample.xlsx")