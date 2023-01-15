from openpyxl import load_workbook
# wb = load_workbook("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample_formula.xlsx")
# ws = wb.active

# # 수식 그대로 가져옴
# for row in ws.values: # ws.values: 모든 셀의 값
#     for cell in row:
#         print(cell)

wb = load_workbook("/Users/gunju/Desktop/self study/python/deep/rpa_basic/1_excel/sample_formula.xlsx", data_only = True)
ws = wb.active

# 수식이 아닌 실제 데이터를 가져옴 / 수식은 Python에서 계산하지 않기 때문에 None으로 나옴
# evaluate 되지 않은 상태의 데이터는 None으로 표시
# 이것을 정상 계산값으로 가져오려면 엑셀 파일을 열고 저장 후 다시 프로그램을 실행시키면 됨
for row in ws.values: # ws.values: 모든 셀의 값
    for cell in row:
        print(cell)